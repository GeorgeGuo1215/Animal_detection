"""
Batch-extract text from all PDFs (and Excel) in 大熊猫/ → RAG/data/raw/*.txt

Usage:
    python extract_pdf.py [--pdf-dir 大熊猫] [--out-dir RAG/data/raw]
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF


_MIN_CHARS_PER_PAGE = 30


def sanitize_filename(name: str) -> str:
    name = re.sub(r"[/\\:*?\"<>|]", "_", name)
    name = re.sub(r"\s+", "_", name.strip())
    if len(name) > 120:
        name = name[:120]
    return name


def extract_pdf(pdf_path: Path) -> tuple[str, dict]:
    doc = fitz.open(str(pdf_path))
    pages: list[str] = []
    empty_pages = 0

    for i, page in enumerate(doc):
        text = page.get_text("text").strip()
        if len(text) < _MIN_CHARS_PER_PAGE:
            empty_pages += 1
        pages.append(f"--- Page {i + 1} of {len(doc)} ---\n\n{text}")

    doc.close()
    full_text = "\n\n".join(pages)
    total_chars = sum(len(p) for p in pages)
    avg_chars = total_chars / max(len(pages), 1)

    stats = {
        "total_pages": len(pages),
        "total_chars": total_chars,
        "avg_chars_per_page": round(avg_chars, 1),
        "empty_pages": empty_pages,
        "is_likely_scanned": avg_chars < 100,
    }
    return full_text, stats


def extract_excel(xlsx_path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"=== Sheet: {sheet_name} ===\n")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells).strip()
            if line:
                parts.append(line)
    wb.close()
    return "\n".join(parts)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract text from PDFs/Excel for RAG ingestion")
    parser.add_argument("--pdf-dir", type=str, default=None,
                        help="Directory containing PDFs (default: 大熊猫/)")
    parser.add_argument("--out-dir", type=str, default=None,
                        help="Output directory for .txt files (default: RAG/data/raw/)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    pdf_dir = Path(args.pdf_dir) if args.pdf_dir else script_dir / "大熊猫"
    out_dir = Path(args.out_dir) if args.out_dir else script_dir / "RAG" / "data" / "raw"

    if not pdf_dir.exists():
        print(f"[Error] PDF directory not found: {pdf_dir}")
        sys.exit(1)

    out_dir.mkdir(parents=True, exist_ok=True)

    pdfs = sorted(pdf_dir.glob("*.pdf"))
    xlsxs = sorted(pdf_dir.glob("*.xlsx"))
    print(f"[Info] Found {len(pdfs)} PDFs + {len(xlsxs)} Excel files in {pdf_dir}")

    scanned_books: list[str] = []

    for i, pdf_path in enumerate(pdfs, 1):
        print(f"\n[{i}/{len(pdfs)}] Processing: {pdf_path.name}")
        try:
            text, stats = extract_pdf(pdf_path)
            print(f"  Pages: {stats['total_pages']}, Chars: {stats['total_chars']}, "
                  f"Avg/page: {stats['avg_chars_per_page']}, Empty pages: {stats['empty_pages']}")

            if stats["is_likely_scanned"]:
                print(f"  ⚠️  Likely scanned (avg {stats['avg_chars_per_page']} chars/page) — saving anyway")
                scanned_books.append(pdf_path.name)

            out_name = sanitize_filename(pdf_path.stem) + ".txt"
            out_path = out_dir / out_name
            out_path.write_text(text, encoding="utf-8")
            print(f"  → Saved: {out_path.name} ({len(text)} chars)")

        except Exception as e:
            print(f"  ✗ Failed: {e}")

    for xlsx_path in xlsxs:
        print(f"\n[Excel] Processing: {xlsx_path.name}")
        try:
            text = extract_excel(xlsx_path)
            out_name = sanitize_filename(xlsx_path.stem) + ".txt"
            out_path = out_dir / out_name
            out_path.write_text(text, encoding="utf-8")
            print(f"  → Saved: {out_path.name} ({len(text)} chars)")
        except Exception as e:
            print(f"  ✗ Failed: {e}")

    print(f"\n{'='*60}")
    print(f"Done. Extracted {len(pdfs)} PDFs + {len(xlsxs)} Excel → {out_dir}")
    if scanned_books:
        print(f"\n⚠️  Likely scanned (low text) — may need OCR:")
        for name in scanned_books:
            print(f"  - {name}")


if __name__ == "__main__":
    main()
